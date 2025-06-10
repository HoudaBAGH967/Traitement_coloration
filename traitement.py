#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, range_boundaries
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import COLOR_INDEX, RGB
import xlsxwriter
import io
import re
import time
import sys
import traceback

# --- CHEMINS DES FICHIERS ---
# Remplacez ces chemins par les vôtres si nécessaire
INPUT_DATA_FILE_PATH = sys.argv[1].strip()
OUTPUT_FILE_PATH = sys.argv[2].strip()
NOM_FICHIER_MOTS_CLES_PATH = sys.argv[3].strip()

# --- CONSTANTES GLOBALES ---
NOM_FICHIER_MOTS_CLES = NOM_FICHIER_MOTS_CLES_PATH # Utilisation du chemin défini ci-dessus
ONGLET_CHAMPS = "Champs"
ONGLET_CONNECTEURS = "Connecteurs"
ONGLET_VARIABLES = "Variables"
ONGLET_TABLES = "Tables"

PROPRIETES_CELLULE_DONNEES_DEFAUT = {
    'border': 1,
    'valign': 'top',
    'align': 'left',
    'text_wrap': True,
}

def lire_mots_depuis_onglet(nom_onglet_a_lire):
    """Lit les mots-clés depuis une colonne d'un onglet spécifié."""
    try:
        tableau_donnees = pd.read_excel(NOM_FICHIER_MOTS_CLES, sheet_name=nom_onglet_a_lire, header=None, usecols=[0], dtype=str)
        if tableau_donnees.empty:
            print(f"INFO : L'onglet '{nom_onglet_a_lire}' dans '{NOM_FICHIER_MOTS_CLES}' est vide.")
            return []
        mots_bruts = tableau_donnees.iloc[:, 0].tolist()
        mots_propres = [str(mot_brut).strip() for mot_brut in mots_bruts if pd.notna(mot_brut) and str(mot_brut).strip()]
        if not mots_propres and not tableau_donnees.empty:
             print(f"INFO : L'onglet '{nom_onglet_a_lire}' ne contient que des valeurs vides après traitement.")
        return mots_propres
    except FileNotFoundError:
        message_erreur = f"ERREUR CRITIQUE : Fichier dictionnaire '{NOM_FICHIER_MOTS_CLES}' non trouvé."
        print(message_erreur)
        print("L'application ne peut pas fonctionner sans ce fichier. Arrêt du script.")
        sys.exit(1)
    except ValueError as ve:
         message_erreur = f"AVERTISSEMENT : Onglet '{nom_onglet_a_lire}' non trouvé dans '{NOM_FICHIER_MOTS_CLES}' ou problème de format. Erreur : {ve}"
         print(message_erreur)
         return []
    except Exception as e:
        message_erreur = f"AVERTISSEMENT : Erreur inattendue lors de la lecture de '{nom_onglet_a_lire}': {e}"
        print(message_erreur)
        return []

# --- CHARGEMENT DES MOTS-CLÉS AU DÉMARRAGE DU SCRIPT ---
print("Chargement du dictionnaire de mots-clés...")
DICTIONNAIRE_MOTS_CLES = {
    'champs': lire_mots_depuis_onglet(ONGLET_CHAMPS),
    'tables': lire_mots_depuis_onglet(ONGLET_TABLES),
    'variables': lire_mots_depuis_onglet(ONGLET_VARIABLES),
    'connecteurs': lire_mots_depuis_onglet(ONGLET_CONNECTEURS)
}
print("Chargement terminé.")

PROPRIETES_SPECIFIQUES_MOTS_CLES = {
    'champs': {'font_color': '#008000', 'bold': True},
    'tables': {'font_color': '#0000FF', 'bold': True},
    'variables': {'font_color': '#FF8C00', 'bold': True},
    'connecteurs': {'font_color': '#FF0000', 'bold': True}
}

def couleur_openpyxl_vers_hex(objet_couleur_openpyxl):
    if objet_couleur_openpyxl is None:
        return None
    valeur_couleur = objet_couleur_openpyxl.value
    if objet_couleur_openpyxl.type == 'rgb':
        return f"#{valeur_couleur[2:]}" if len(valeur_couleur) == 8 else f"#{valeur_couleur}"
    elif objet_couleur_openpyxl.type == 'indexed':
        if 0 <= valeur_couleur < len(COLOR_INDEX):
            couleur_indexee_rgb = COLOR_INDEX[valeur_couleur]
            return f"#{couleur_indexee_rgb[2:]}" if len(couleur_indexee_rgb) == 8 else f"#{couleur_indexee_rgb}"
        return None
    elif objet_couleur_openpyxl.type == 'theme':
        return None
    return None

def proprietes_format_xlsxwriter_depuis_cellule_openpyxl(cellule_openpyxl, proprietes_defaut_base=None):
    proprietes_format = proprietes_defaut_base.copy() if proprietes_defaut_base else {}

    if cellule_openpyxl.has_style and cellule_openpyxl.font:
        police = cellule_openpyxl.font
        if police.name: proprietes_format['font_name'] = police.name
        if police.sz: proprietes_format['font_size'] = police.sz
        if police.bold: proprietes_format['bold'] = True
        if police.italic: proprietes_format['italic'] = True
        if police.underline and police.underline != 'none': proprietes_format['underline'] = True
        if police.strike: proprietes_format['font_strikeout'] = True
        if police.color and police.color.value:
            couleur_hex = couleur_openpyxl_vers_hex(police.color)
            if couleur_hex: proprietes_format['font_color'] = couleur_hex

    if cellule_openpyxl.has_style and cellule_openpyxl.fill and cellule_openpyxl.fill.fgColor:
        if cellule_openpyxl.fill.fill_type in ['solid', 'patternGrid', 'patternGray125', 'patternGray0625', 'patternLightDown',
                                   'patternLightGray', 'patternLightGrid', 'patternLightHorizontal',
                                   'patternLightTrellis', 'patternLightUp', 'patternLightVertical',
                                   'patternMediumGray', 'patternDarkDown', 'patternDarkGray',
                                   'patternDarkGrid', 'patternDarkHorizontal', 'patternDarkTrellis',
                                   'patternDarkUp', 'patternDarkVertical']:
            couleur_remplissage = couleur_openpyxl_vers_hex(cellule_openpyxl.fill.fgColor)
            if couleur_remplissage:
                proprietes_format['bg_color'] = couleur_remplissage
                if cellule_openpyxl.fill.fill_type != 'solid':
                    pass

    if cellule_openpyxl.has_style and cellule_openpyxl.border:
        bordure = cellule_openpyxl.border
        correspondance_style_bordure = {
            'thin': 1, 'medium': 2, 'thick': 5, 'double': 6,
            'dotted': 3, 'dashed': 4, 'hair': 7,
        }
        if bordure.left and bordure.left.style and bordure.left.style != 'none':
            proprietes_format['left'] = correspondance_style_bordure.get(bordure.left.style, 1)
            if bordure.left.color: proprietes_format['left_color'] = couleur_openpyxl_vers_hex(bordure.left.color)
        if bordure.right and bordure.right.style and bordure.right.style != 'none':
            proprietes_format['right'] = correspondance_style_bordure.get(bordure.right.style, 1)
            if bordure.right.color: proprietes_format['right_color'] = couleur_openpyxl_vers_hex(bordure.right.color)
        if bordure.top and bordure.top.style and bordure.top.style != 'none':
            proprietes_format['top'] = correspondance_style_bordure.get(bordure.top.style, 1)
            if bordure.top.color: proprietes_format['top_color'] = couleur_openpyxl_vers_hex(bordure.top.color)
        if bordure.bottom and bordure.bottom.style and bordure.bottom.style != 'none':
            proprietes_format['bottom'] = correspondance_style_bordure.get(bordure.bottom.style, 1)
            if bordure.bottom.color: proprietes_format['bottom_color'] = couleur_openpyxl_vers_hex(bordure.bottom.color)

    if cellule_openpyxl.has_style and cellule_openpyxl.alignment:
        alignement = cellule_openpyxl.alignment
        if alignement.horizontal: proprietes_format['align'] = alignement.horizontal
        if alignement.vertical: proprietes_format['valign'] = alignement.vertical
        if alignement.wrap_text is not None: proprietes_format['text_wrap'] = alignement.wrap_text
        if alignement.shrink_to_fit is not None: proprietes_format['shrink'] = alignement.shrink_to_fit
        if alignement.indent: proprietes_format['indent'] = alignement.indent
        if alignement.text_rotation: proprietes_format['rotation'] = alignement.text_rotation

    if cellule_openpyxl.has_style and cellule_openpyxl.number_format and cellule_openpyxl.number_format != 'General':
        proprietes_format['num_format'] = cellule_openpyxl.number_format

    return proprietes_format

def obtenir_ou_creer_format_xlsxwriter_cache(classeur_sortie_xlsx, cache_formats, dictionnaire_proprietes_format):
    cle_format = frozenset(sorted(dictionnaire_proprietes_format.items()))
    if cle_format not in cache_formats:
        cache_formats[cle_format] = classeur_sortie_xlsx.add_format(dictionnaire_proprietes_format)
    return cache_formats[cle_format]

def detecter_mots_cles_texte(texte_original):
    if not isinstance(texte_original, str): return []
    mots_cles_trouves = []
    texte_majuscule = texte_original.upper()
    for type_mot_cle, liste_mots_type in DICTIONNAIRE_MOTS_CLES.items():
        if not liste_mots_type: continue
        for mot_reference in liste_mots_type:
            mot_reference_majuscule = mot_reference.upper()
            if not mot_reference_majuscule: continue
            if ' ' in mot_reference_majuscule:
                index_recherche = 0
                while True:
                    index_trouve = texte_majuscule.find(mot_reference_majuscule, index_recherche)
                    if index_trouve == -1: break
                    limites_valides = True
                    if index_trouve > 0 and texte_majuscule[index_trouve-1].isalnum() and mot_reference_majuscule[0].isalnum():
                        limites_valides = False
                    if limites_valides and (index_trouve + len(mot_reference_majuscule) < len(texte_majuscule)) and \
                       texte_majuscule[index_trouve+len(mot_reference_majuscule)].isalnum() and mot_reference_majuscule[-1].isalnum():
                        limites_valides = False
                    if limites_valides:
                        mots_cles_trouves.append((type_mot_cle, mot_reference, index_trouve, index_trouve + len(mot_reference_majuscule)))
                    index_recherche = index_trouve + len(mot_reference_majuscule)
            else:
                motif_regex = r'(?<!\w)' + re.escape(mot_reference_majuscule) + r'(?!\w)'
                for correspondance_regex in re.finditer(motif_regex, texte_majuscule):
                    mots_cles_trouves.append((type_mot_cle, mot_reference, correspondance_regex.start(), correspondance_regex.end()))
    mots_cles_trouves.sort(key=lambda x: (x[2], -(x[3] - x[2])))
    mots_cles_filtres = []
    derniere_position_fin = -1
    for mot_cle_detecte in mots_cles_trouves:
        if mot_cle_detecte[2] >= derniere_position_fin:
            mots_cles_filtres.append(mot_cle_detecte)
            derniere_position_fin = mot_cle_detecte[3]
    return mots_cles_filtres

def appliquer_formatage_enrichi_cellule(onglet_sortie_xlsx, ligne, colonne, contenu_cellule, mots_cles_detectes, formats_par_type_mot_cle, format_cellule_defaut):
    if not isinstance(contenu_cellule, str) or not contenu_cellule.strip():
        onglet_sortie_xlsx.write_string(ligne, colonne, contenu_cellule, format_cellule_defaut)
        return
    if not mots_cles_detectes:
        onglet_sortie_xlsx.write_string(ligne, colonne, contenu_cellule, format_cellule_defaut)
        return
    elements_texte_enrichi = []
    index_caractere_actuel = 0
    for type_mot_cle, _, debut_mot, fin_mot in mots_cles_detectes:
        if debut_mot > index_caractere_actuel:
            elements_texte_enrichi.append(contenu_cellule[index_caractere_actuel:debut_mot])
        format_superposition_mot_cle = formats_par_type_mot_cle[type_mot_cle]
        elements_texte_enrichi.append(format_superposition_mot_cle)
        elements_texte_enrichi.append(contenu_cellule[debut_mot:fin_mot])
        index_caractere_actuel = fin_mot
    if index_caractere_actuel < len(contenu_cellule):
        elements_texte_enrichi.append(contenu_cellule[index_caractere_actuel:])
    if elements_texte_enrichi:
        onglet_sortie_xlsx.write_rich_string(ligne, colonne, *elements_texte_enrichi, format_cellule_defaut)
    else:
        onglet_sortie_xlsx.write_string(ligne, colonne, contenu_cellule, format_cellule_defaut)

def traiter_fichier_excel(chemin_fichier_entree):
    """
    Charge un fichier Excel, applique la coloration des mots-clés et le formatage,
    et retourne le fichier modifié sous forme de flux d'octets.
    """
    classeur_origine_openpyxl = load_workbook(chemin_fichier_entree, data_only=True)
    flux_sortie_octets = io.BytesIO()

    with xlsxwriter.Workbook(flux_sortie_octets, {'in_memory': True, 'strings_to_urls': False, 'remove_timezone': True}) as classeur_sortie_xlsx:
        cache_general_formats_xlsx = {}
        format_cellules_donnees_defaut = obtenir_ou_creer_format_xlsxwriter_cache(
            classeur_sortie_xlsx, cache_general_formats_xlsx, PROPRIETES_CELLULE_DONNEES_DEFAUT
        )

        formats_mots_cles_complets = {}
        formats_mots_cles_superposition = {}

        for type_mot_cle, proprietes_specifiques in PROPRIETES_SPECIFIQUES_MOTS_CLES.items():
            proprietes_mot_cle_complet = PROPRIETES_CELLULE_DONNEES_DEFAUT.copy()
            proprietes_mot_cle_complet.update(proprietes_specifiques)
            formats_mots_cles_complets[type_mot_cle] = obtenir_ou_creer_format_xlsxwriter_cache(
                classeur_sortie_xlsx, cache_general_formats_xlsx, proprietes_mot_cle_complet
            )
            formats_mots_cles_superposition[type_mot_cle] = obtenir_ou_creer_format_xlsxwriter_cache(
                classeur_sortie_xlsx, cache_general_formats_xlsx, proprietes_specifiques.copy()
            )

        for nom_onglet_origine in classeur_origine_openpyxl.sheetnames:
            onglet_origine_openpyxl = classeur_origine_openpyxl[nom_onglet_origine]
            onglet_sortie_xlsx = classeur_sortie_xlsx.add_worksheet(nom_onglet_origine)

            if onglet_origine_openpyxl.auto_filter and onglet_origine_openpyxl.auto_filter.ref:
                try:
                    col_min, ligne_min, col_max, ligne_max = range_boundaries(onglet_origine_openpyxl.auto_filter.ref)
                    onglet_sortie_xlsx.autofilter(ligne_min - 1, col_min - 1, ligne_max - 1, col_max - 1)
                except Exception as erreur_autofiltre:
                    print(f"AVERTISSEMENT: Copie autofiltre échouée pour '{nom_onglet_origine}': {erreur_autofiltre}")

            largeur_defaut_retour_ligne = 15.0
            colonnes_largeur_perso_indices = set()
            if hasattr(onglet_origine_openpyxl, 'column_dimensions'):
                for lettre_colonne, dimension_colonne in onglet_origine_openpyxl.column_dimensions.items():
                    try:
                        num_colonne_base0 = column_index_from_string(lettre_colonne) - 1
                        if dimension_colonne.customWidth and dimension_colonne.width is not None and dimension_colonne.width > 0:
                            onglet_sortie_xlsx.set_column(num_colonne_base0, num_colonne_base0, dimension_colonne.width)
                            colonnes_largeur_perso_indices.add(num_colonne_base0)
                    except Exception as erreur_largeur_colonne:
                        print(f"Avertissement : Largeur colonne '{lettre_colonne}': {erreur_largeur_colonne}")

            max_col_origine_base0 = (onglet_origine_openpyxl.max_column -1) if onglet_origine_openpyxl.max_column else -1
            for index_col_base0 in range(max_col_origine_base0 + 1):
                if index_col_base0 not in colonnes_largeur_perso_indices:
                    onglet_sortie_xlsx.set_column(index_col_base0, index_col_base0, largeur_defaut_retour_ligne)

            for num_ligne, ligne_origine_openpyxl in enumerate(onglet_origine_openpyxl.iter_rows()):
                cle_dimension_ligne = num_ligne + 1
                if cle_dimension_ligne in onglet_origine_openpyxl.row_dimensions:
                    dimension_ligne = onglet_origine_openpyxl.row_dimensions[cle_dimension_ligne]
                    if dimension_ligne.customHeight and dimension_ligne.height is not None:
                         onglet_sortie_xlsx.set_row(num_ligne, dimension_ligne.height)
                    else:
                         onglet_sortie_xlsx.set_row(num_ligne, None)
                else:
                    onglet_sortie_xlsx.set_row(num_ligne, None)

                for num_colonne, cellule_origine_openpyxl in enumerate(ligne_origine_openpyxl):
                    contenu_cellule = cellule_origine_openpyxl.value

                    if num_ligne == 0:
                        proprietes_base_entete = {'text_wrap': True, 'valign': 'top', 'align': 'left', 'border':1}
                        proprietes_xlsxwriter_cellule_entete = proprietes_format_xlsxwriter_depuis_cellule_openpyxl(
                            cellule_origine_openpyxl,
                            proprietes_base_entete
                        )
                        if not proprietes_xlsxwriter_cellule_entete or not any(k not in proprietes_base_entete for k in proprietes_xlsxwriter_cellule_entete):
                            format_final_entete = obtenir_ou_creer_format_xlsxwriter_cache(
                                classeur_sortie_xlsx, cache_general_formats_xlsx, proprietes_base_entete
                            )
                        else:
                            format_final_entete = obtenir_ou_creer_format_xlsxwriter_cache(
                                classeur_sortie_xlsx, cache_general_formats_xlsx, proprietes_xlsxwriter_cellule_entete
                            )

                        if contenu_cellule is None:
                            onglet_sortie_xlsx.write_blank(num_ligne, num_colonne, None, format_final_entete)
                        elif isinstance(contenu_cellule, (int, float)):
                            onglet_sortie_xlsx.write_number(num_ligne, num_colonne, contenu_cellule, format_final_entete)
                        elif isinstance(contenu_cellule, bool):
                            onglet_sortie_xlsx.write_boolean(num_ligne, num_colonne, contenu_cellule, format_final_entete)
                        else:
                            onglet_sortie_xlsx.write_string(num_ligne, num_colonne, str(contenu_cellule), format_final_entete)
                        continue

                    format_base_cellule_donnees = format_cellules_donnees_defaut

                    if contenu_cellule is None:
                        onglet_sortie_xlsx.write_blank(num_ligne, num_colonne, None, format_base_cellule_donnees)
                        continue

                    if isinstance(contenu_cellule, str) and contenu_cellule.strip():
                        type_mot_cle_correspondance_exacte = None
                        contenu_cellule_majuscule = contenu_cellule.upper()
                        for type_mot_cle, mots_du_type in DICTIONNAIRE_MOTS_CLES.items():
                            if contenu_cellule_majuscule in [m.upper() for m in mots_du_type if m]:
                                type_mot_cle_correspondance_exacte = type_mot_cle
                                break

                        if type_mot_cle_correspondance_exacte:
                            format_mot_cle_complet = formats_mots_cles_complets[type_mot_cle_correspondance_exacte]
                            onglet_sortie_xlsx.write_string(num_ligne, num_colonne, contenu_cellule, format_mot_cle_complet)
                        else:
                            mots_cles_detectes_cellule = detecter_mots_cles_texte(contenu_cellule)
                            if mots_cles_detectes_cellule:
                                appliquer_formatage_enrichi_cellule(
                                    onglet_sortie_xlsx, num_ligne, num_colonne, contenu_cellule,
                                    mots_cles_detectes_cellule, formats_mots_cles_superposition, format_base_cellule_donnees
                                )
                            else:
                                onglet_sortie_xlsx.write_string(num_ligne, num_colonne, contenu_cellule, format_base_cellule_donnees)
                    else:
                        if isinstance(contenu_cellule, (int, float)):
                            onglet_sortie_xlsx.write_number(num_ligne, num_colonne, contenu_cellule, format_base_cellule_donnees)
                        elif isinstance(contenu_cellule, bool):
                            onglet_sortie_xlsx.write_boolean(num_ligne, num_colonne, contenu_cellule, format_base_cellule_donnees)
                        else:
                            onglet_sortie_xlsx.write(num_ligne, num_colonne, str(contenu_cellule) if contenu_cellule is not None else "", format_base_cellule_donnees)

            if hasattr(onglet_origine_openpyxl, 'merged_cells') and hasattr(onglet_origine_openpyxl.merged_cells, 'ranges'):
                for plage_cellule_fusionnee in onglet_origine_openpyxl.merged_cells.ranges:
                    try:
                        onglet_sortie_xlsx.merge_range(
                            plage_cellule_fusionnee.min_row - 1, plage_cellule_fusionnee.min_col - 1,
                            plage_cellule_fusionnee.max_row - 1, plage_cellule_fusionnee.max_col - 1, "", None
                        )
                    except Exception as erreur_fusion:
                        print(f"Avertissement: Fusion plage échouée {str(plage_cellule_fusionnee)}: {erreur_fusion}")

    flux_sortie_octets.seek(0)
    return flux_sortie_octets

def main():
    """Fonction principale du script."""
    print("="*50)
    print("Démarrage du script de coloration Excel")
    print("="*50)
    print(f"Fichier d'entrée : {INPUT_DATA_FILE_PATH}")
    print(f"Fichier de sortie : {OUTPUT_FILE_PATH}")
    print(f"Fichier dictionnaire : {NOM_FICHIER_MOTS_CLES}")
    print("-" * 50)

    nombre_total_mots_cles = sum(len(valeur_liste_mots) for valeur_liste_mots in DICTIONNAIRE_MOTS_CLES.values() if isinstance(valeur_liste_mots, list))
    if nombre_total_mots_cles == 0:
        print("ERREUR CRITIQUE : Aucun mot-clé n'a été chargé.")
        print("Veuillez vérifier que le fichier dictionnaire et ses feuilles ne sont pas vides.")
        sys.exit(1)

    print(f"{nombre_total_mots_cles} mots-clés chargés au total.")

    try:
        temps_debut_traitement = time.time()
        print("\nLancement du traitement du fichier Excel (formatage et coloration)...")

        # Appel de la fonction de traitement avec le chemin du fichier d'entrée
        flux_sortie_traite = traiter_fichier_excel(INPUT_DATA_FILE_PATH)

        # Écriture du fichier de sortie
        with open(OUTPUT_FILE_PATH, 'wb') as f_out:
            f_out.write(flux_sortie_traite.getvalue())

        temps_fin_traitement = time.time()
        duree_traitement = temps_fin_traitement - temps_debut_traitement

        print(f"\n---> Traitement terminé avec succès en {duree_traitement:.2f} secondes !")
        print(f"---> Le fichier traité a été sauvegardé ici : {OUTPUT_FILE_PATH}")

    except FileNotFoundError:
        print(f"\nERREUR: Le fichier d'entrée '{INPUT_DATA_FILE_PATH}' n'a pas été trouvé.")
        print("Veuillez vérifier le chemin d'accès et le nom du fichier.")
    except Exception as erreur_generale:
        print(f"\nUne erreur inattendue est survenue lors du traitement : {erreur_generale}")
        print("Trace de l'erreur :")
        traceback.print_exc()

if __name__ == "__main__":
    main()

